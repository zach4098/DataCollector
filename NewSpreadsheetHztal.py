from openpyxl import Workbook, load_workbook

wb = Workbook()
workbookName = input("Enter Spreadsheet Name: ")
if workbookName == "":
    workbookName = "data"
wb.save("DataSpread/{}.xlsx".format(workbookName))
ws = wb.active
ws.column_dimensions['A'].width = 30
ws.cell(1, 1).value = "File Name"
ws.cell(2, 1).value = "Time Initial"
ws.cell(3, 1).value = "Time Final"
ws.cell(4, 1).value = "Time Total"
ws.cell(5, 1).value = "Total Vehicles"
ws.cell(6, 1).value = "Vehicles from Left"
ws.cell(7, 1).value = "Vehicles from Right"
ws.cell(8, 1).value = "Vehicles from N/A"
ws.cell(9, 1).value = "Total Hours"
ws.cell(10, 1).value = "Vehicles per Hour"
ws.cell(11, 1).value = "Vehicles per Hour from Left"
ws.cell(12, 1).value = "Vehicles per Hour from Right"
ws.cell(13, 1).value = "Morning Peak Start Time"
ws.cell(14, 1).value = "Morning Peak End Time"
ws.cell(15, 1).value = "Morning Peak Total Vehicles"
ws.cell(16, 1).value = "Morning Peak Total Hours"
ws.cell(17, 1).value = "Morning Peak Vehicles per Hour"
ws.cell(18, 1).value = "Night Peak Start Time"
ws.cell(19, 1).value = "Night Peak End Time"
ws.cell(20, 1).value = "Night Peak Total Vehicles"
ws.cell(21, 1).value = "Night Peak Total Hours"
ws.cell(22, 1).value = "Night Peak Vehicles per Hour"

wb.save("DataSpread/{}.xlsx".format(workbookName))