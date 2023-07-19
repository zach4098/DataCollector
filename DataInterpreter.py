import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

def ReadTime(text):
    time = text.split('-')[1].split(' ')[3].split(':')
    return time
def TimeDifference(timeInit, timeFinal, vehicleCount, leftCount, rightCount):
    if timeInit == "N/A":
        differenceInHours = differenceInMinutes = differenceInSeconds = totalHours = VperHour = VperHrL = VperHrR = "N/A"
    else:
        initHour = int(timeInit[0])
        initMinute = int(timeInit[1])
        initSecond = int(timeInit[2])

        finalHour = int(timeFinal[0])
        finalMinute = int(timeFinal[1])
        finalSecond = int(timeFinal[2])

        differenceInHours = finalHour-initHour
        differenceInMinutes = finalMinute - initMinute
        differenceInSeconds = finalSecond - initSecond

        if differenceInMinutes < 0:
            differenceInHours -= 1
            differenceInMinutes += 60
        if differenceInSeconds < 0:
            differenceInMinutes -= 1
            differenceInSeconds += 60
        
        totalHours = differenceInHours + (differenceInMinutes/60) + (differenceInSeconds/3600)
        totalHours = round(totalHours, 5)
        
        VperHour = round(vehicleCount/totalHours, 2)
        VperHrL = round(leftCount/totalHours, 2)
        VperHrR = round(rightCount/totalHours, 2)


    return differenceInHours, differenceInMinutes, differenceInSeconds, totalHours, VperHour, VperHrL, VperHrR
def CountLeftRight(List):
    leftCount = 0
    rightCount = 0
    nullCount = 0
    for item in List:
        direction = item.split('-')[2]
        if direction == "left":
            leftCount += 1
        elif direction == "right":
            rightCount += 1
        else:
            nullCount += 1
    return leftCount, rightCount, nullCount

MorningPeakHours = [8, 10]
NightPeakHours = [17, 21]

def PeakHours(timeInit, timeFinal, folder, file):
    peakHoursM = False
    peakHoursN = False
    if int(timeInit[0]) <= MorningPeakHours[0]:
        peakHoursM = True
    if int(timeFinal[0]) >= NightPeakHours[0]:
        peakHoursN = True
    if peakHoursM:
        count = 0
        hoursFound = False
        with open("{}/{}".format(folder, file), "r") as f:
            lines = f.readlines()
        for item in lines:
            item = item.replace("\n", "")
            item = item.replace("Vehicle ", "")
            lines[count] = item
            count += 1
        count = 0
        start = False
        while not start:
            time = ReadTime(lines[count])
            if int(time[0]) < MorningPeakHours[0]:
                count += 1
                print(count)
            else:
                start = True
        vehicleMCount = 0
        peakHoursMVehicles = []
        while not hoursFound:
            if count < len(lines):
                time = ReadTime(lines[count])
                if vehicleMCount == 0:
                    initTimeM = time
                if MorningPeakHours[0] <= int(time[0]) <= MorningPeakHours[1]:
                    peakHoursMVehicles.append(time)
                    count += 1
                    vehicleMCount += 1
                if int(time[0]) > MorningPeakHours[1]:
                    finalTimeM = ReadTime(lines[count - 1])
                    hoursFound = True
            else:
                finalTimeM = ReadTime(lines[count - 1])
                hoursFound = True
    if peakHoursN:
        count = 0
        hoursFound = False
        with open("{}/{}".format(folder, file), "r") as f:
            lines = f.readlines()
        for item in lines:
            item = item.replace("\n", "")
            item = item.replace("Vehicle ", "")
            lines[count] = item
            count += 1
        count = 0
        start = False
        while not start:
            time = ReadTime(lines[count])
            if int(time[0]) < NightPeakHours[0]:
                count += 1
            else:
                start = True
        vehicleNCount = 0
        peakHoursNVehicles = []
        while not hoursFound:
            if count < len(lines):
                time = ReadTime(lines[count])
                if vehicleNCount == 0:
                    initTimeN = time
                if NightPeakHours[0] <= int(time[0]) <= NightPeakHours[1]:
                    peakHoursNVehicles.append(time)
                    count += 1
                    vehicleNCount += 1
                if int(time[0]) > NightPeakHours[1]:
                    finalTimeN = ReadTime(lines[count - 1])
                    hoursFound = True
            else:
                finalTimeN = ReadTime(lines[count - 1])
                hoursFound = True
    if not peakHoursM:
        vehicleMCount = "N/A"
        initTimeM = "N/A"
        finalTimeM = "N/A"
    if not peakHoursN:
        vehicleNCount = "N/A"
        initTimeN = "N/A"
        finalTimeN = "N/A"
    return vehicleMCount, initTimeM, finalTimeM, vehicleNCount, initTimeN, finalTimeN
    


def ReadFile(folder, file):
    
    with open("{}/{}".format(folder, file), "r") as f:
            lines = f.readlines()
    count = 0
    for item in lines:
        item = item.replace("\n", "")
        item = item.replace("Vehicle ", "")
        lines[count] = item
        count += 1
    totalVehicles = count
    totalLeft, totalRight, totalNull = CountLeftRight(lines)
    lineInit = lines[0]
    lineFinal = lines[count - 1]
    timeInit = ReadTime(lineInit)
    timeFinal = ReadTime(lineFinal)

    differenceHour, differenceMinute, differenceSecond, totalHours, vPerHour, vPerHourL, vPerHourR = TimeDifference(timeInit, timeFinal, totalVehicles, totalLeft, totalRight)

    vehicleMCount, initTimeM, finalTimeM, vehicleNCount, initTimeN, finalTimeN = PeakHours(timeInit, timeFinal, folder, file)

    MdifferenceHour, MdifferenceMinute, MdifferenceSecond, MtotalHours, MvPerHour, MvPerHourL, MvPerHourR = TimeDifference(initTimeM, finalTimeM, vehicleMCount, 0, 0)
    NdifferenceHour, NdifferenceMinute, NdifferenceSecond, NtotalHours, NvPerHour, NvPerHourL, NvPerHourR = TimeDifference(initTimeN, finalTimeN, vehicleNCount, 0, 0)

    wb = load_workbook("DataSpread/data.xlsx")
    ws = wb.active
    currentRow = 2
    openRow = False
    while not openRow:
        if str(ws.cell(currentRow, 1).value) == "None":
            openRow = True
        else:
            currentRow += 1
    ws.cell(currentRow, 1).value = file
    ws.cell(currentRow, 2).value = "{}:{}:{}".format(timeInit[0], timeInit[1], timeInit[2])
    ws.cell(currentRow, 3).value = "{}:{}:{}".format(timeFinal[0], timeFinal[1], timeFinal[2])
    ws.cell(currentRow, 4).value = "{}:{}:{}".format(differenceHour, differenceMinute, differenceSecond)
    ws.cell(currentRow, 5).value = totalVehicles
    ws.cell(currentRow, 6).value = totalLeft
    ws.cell(currentRow, 7).value = totalRight
    ws.cell(currentRow, 8).value = totalNull
    ws.cell(currentRow, 9).value = totalHours
    ws.cell(currentRow, 10).value = vPerHour
    ws.cell(currentRow, 11).value = vPerHourL
    ws.cell(currentRow, 12).value = vPerHourR
    ws.cell(currentRow, 13).value = "{}:{}:{}".format(initTimeM[0], initTimeM[1], initTimeM[2])
    ws.cell(currentRow, 14).value = "{}:{}:{}".format(finalTimeM[0], finalTimeM[1], finalTimeM[2])
    ws.cell(currentRow, 15).value = vehicleMCount
    ws.cell(currentRow, 16).value = MtotalHours
    ws.cell(currentRow, 17).value = MvPerHour
    ws.cell(currentRow, 18).value = "{}:{}:{}".format(initTimeN[0], initTimeN[1], initTimeN[2])
    ws.cell(currentRow, 19).value = "{}:{}:{}".format(finalTimeN[0], finalTimeN[1], finalTimeN[2])
    ws.cell(currentRow, 20).value = vehicleNCount
    ws.cell(currentRow, 21).value = NtotalHours
    ws.cell(currentRow, 22).value = NvPerHour


    wb.save("DataSpread/data.xlsx")
folder = input("Select Folder: ")
collections = os.listdir("{}/".format(folder))

msg = "Select File to Read:\n"
count = 0
for item in collections:
    msg = msg + "{}: {}".format(count, str(item)) + "\n"
    count += 1
print(msg)

fileInput = input("Select File: ")

if fileInput == "all":
    count = 0
    for i in collections:
        ReadFile(folder, collections[count])
        count += 1
    print("Added a total of {} datasets!".format(count))
else:
    file = collections[int(fileInput)]
    ReadFile(folder, file)
    print("Done!")